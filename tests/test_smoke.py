"""Offline smoke tests for reddit-to-spreadsheet.

Every test here runs fully offline -- no network is ever touched. The
collector's ``RTS_FAKE`` hook lets us exercise the whole stack (Flask route
-> collector -> spreadsheet -> .xlsx bytes) deterministically.

Run from the repo root::

    pytest -q
"""

from __future__ import annotations

import io
import json
import os
import sys
import zipfile

import openpyxl
import pytest
import zstandard

# Make the repo root importable no matter where pytest is invoked from.
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import app as app_module  # noqa: E402
import bundle  # noqa: E402
import collector  # noqa: E402
import spreadsheet  # noqa: E402

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
ZIP_MIME = "application/zip"


def _read_ndjson_zst(raw: bytes) -> list[dict]:
    text = zstandard.ZstdDecompressor().decompress(raw).decode("utf-8")
    return [json.loads(line) for line in text.splitlines() if line]


# --------------------------------------------------------------------------- #
# Fixtures / helpers
# --------------------------------------------------------------------------- #
def _stub_posts():
    return [
        {
            "kind": "post", "subreddit": "endometriosis", "id": "p1",
            "created_utc": 1_700_000_000, "author": "alice",
            "title": "First post", "selftext": "hello world",
            "score": 10, "num_comments": 2,
            "permalink": "https://reddit.com/r/endometriosis/p1",
        },
        {
            "kind": "post", "subreddit": "endometriosis", "id": "p2",
            "created_utc": 1_700_100_000, "author": None,
            "title": "Second post", "selftext": "more text here",
            "score": 5, "num_comments": 0,
            "permalink": "https://reddit.com/r/endometriosis/p2",
        },
    ]


def _stub_comments():
    return [
        {
            "kind": "comment", "subreddit": "endometriosis", "id": "c1",
            "created_utc": 1_700_000_500, "author": "bob", "body": "nice",
            "score": 3, "link_id": "t3_p1", "parent_id": "t3_p1",
            "permalink": "https://reddit.com/r/endometriosis/p1/c1",
        },
        {
            "kind": "comment", "subreddit": "endometriosis", "id": "c2",
            "created_utc": 1_700_000_600, "author": None, "body": "thanks",
            "score": 1, "link_id": "t3_p1", "parent_id": "t1_c1",
            "permalink": "https://reddit.com/r/endometriosis/p1/c2",
        },
        {
            "kind": "comment", "subreddit": "endometriosis", "id": "c3",
            "created_utc": 1_700_100_500, "author": "carol", "body": "hi",
            "score": 0, "link_id": "t3_p2", "parent_id": "t3_p2",
            "permalink": "https://reddit.com/r/endometriosis/p2/c3",
        },
    ]


# --------------------------------------------------------------------------- #
# spreadsheet.build_workbook_bytes
# --------------------------------------------------------------------------- #
def test_build_workbook_bytes_sheets_and_rows():
    posts = _stub_posts()
    comments = _stub_comments()

    raw = spreadsheet.build_workbook_bytes(posts, comments)
    assert isinstance(raw, (bytes, bytearray)) and len(raw) > 0

    wb = openpyxl.load_workbook(io.BytesIO(raw))
    names = wb.sheetnames
    assert "Posts" in names
    assert "Comments" in names
    assert "Summary" in names

    # Header row + one row per record.
    assert wb["Posts"].max_row == len(posts) + 1
    assert wb["Comments"].max_row == len(comments) + 1

    # Summary has a header, at least the one subreddit row + a TOTAL row.
    assert wb["Summary"].max_row >= 3


def test_build_workbook_bytes_omits_empty_comments_sheet():
    # Per the contract the Comments sheet is only present when there are
    # comments; Posts + Summary must always exist.
    raw = spreadsheet.build_workbook_bytes(_stub_posts(), [])
    wb = openpyxl.load_workbook(io.BytesIO(raw))
    assert "Posts" in wb.sheetnames
    assert "Summary" in wb.sheetnames
    assert "Comments" not in wb.sheetnames


# --------------------------------------------------------------------------- #
# collector RTS_FAKE hook (offline)
# --------------------------------------------------------------------------- #
def test_collector_fake_hook_is_offline(monkeypatch):
    monkeypatch.setenv("RTS_FAKE", "1")

    result = collector.collect(["endometriosis"], 1_700_000_000, 1_700_200_000)

    assert set(result.keys()) >= {"posts", "comments", "errors"}
    posts = result["posts"]
    comments = result["comments"]

    assert len(posts) > 0
    assert len(comments) > 0
    assert all(p["kind"] == "post" for p in posts)
    assert all(c["kind"] == "comment" for c in comments)
    assert all(p["subreddit"] == "endometriosis" for p in posts)
    # Synthetic data must satisfy the Post schema well enough to spreadsheet.
    for p in posts:
        assert {"id", "created_utc", "title", "score"} <= set(p.keys())


def test_collector_fake_hook_multiple_subs(monkeypatch):
    monkeypatch.setenv("RTS_FAKE", "1")
    result = collector.collect(["endometriosis", "PCOS"], 0, 2_000_000_000)
    subs = {p["subreddit"] for p in result["posts"]}
    assert subs == {"endometriosis", "PCOS"}


# --------------------------------------------------------------------------- #
# app POST /api/collect end-to-end (offline via RTS_FAKE)
# --------------------------------------------------------------------------- #
@pytest.fixture()
def client():
    app_module.app.config.update(TESTING=True)
    return app_module.app.test_client()


def test_collect_endpoint_returns_valid_zip_bundle(client, monkeypatch):
    monkeypatch.setenv("RTS_FAKE", "1")

    resp = client.post(
        "/api/collect",
        json={
            "subreddits": ["endometriosis"],
            "start_date": "2023-11-01",
            "end_date": "2023-11-30",
            "include_comments": True,
            "exclude_usernames": False,
            "max_posts_per_sub": 50,
            "max_comments_per_sub": 100,
        },
    )

    assert resp.status_code == 200
    assert resp.mimetype == ZIP_MIME
    assert "attachment" in resp.headers.get("Content-Disposition", "")
    assert resp.headers["Content-Disposition"].endswith('.zip"')

    zf = zipfile.ZipFile(io.BytesIO(resp.data))
    names = zf.namelist()

    # The workbook inside the bundle must reopen with the expected sheets.
    xlsx_names = [n for n in names if n.endswith(".xlsx")]
    assert len(xlsx_names) == 1
    wb = openpyxl.load_workbook(io.BytesIO(zf.read(xlsx_names[0])))
    assert "Posts" in wb.sheetnames
    assert "Summary" in wb.sheetnames

    # Raw .ndjson.zst dumps: line counts must equal the reported counts.
    assert "raw/endometriosis_posts.ndjson.zst" in names
    assert "raw/endometriosis_comments.ndjson.zst" in names
    posts = _read_ndjson_zst(zf.read("raw/endometriosis_posts.ndjson.zst"))
    comments = _read_ndjson_zst(zf.read("raw/endometriosis_comments.ndjson.zst"))
    assert len(posts) == int(resp.headers["X-Collect-Posts"]) > 0
    assert len(comments) == int(resp.headers["X-Collect-Comments"]) > 0
    assert all(p["kind"] == "post" and p["subreddit"] == "endometriosis" for p in posts)
    assert all("body" in c for c in comments)


def test_collect_endpoint_topic_filter(client, monkeypatch):
    # Fake posts are titled "[sub] Synthetic post #N"; comments say
    # "Fake comment N on a post ...". A topic of "synthetic" (fallback keyword
    # expansion, since RTS_FAKE skips Mercury) must keep posts but no comments.
    monkeypatch.setenv("RTS_FAKE", "1")

    resp = client.post(
        "/api/collect",
        json={
            "subreddits": ["endometriosis"],
            "start_date": "2023-11-01",
            "end_date": "2023-11-30",
            "topic": "synthetic",
        },
    )

    assert resp.status_code == 200
    assert int(resp.headers["X-Collect-Posts"]) > 0
    assert int(resp.headers["X-Collect-Comments"]) == 0
    assert "synthetic" in resp.headers.get("X-Collect-Keywords", "")

    # No comment matched -> no comments dump in the bundle.
    names = zipfile.ZipFile(io.BytesIO(resp.data)).namelist()
    assert "raw/endometriosis_posts.ndjson.zst" in names
    assert "raw/endometriosis_comments.ndjson.zst" not in names


def test_heuristic_dedicated():
    subs = ["BreastCancer", "breastcancerawareness", "cancer", "oncology", "health"]
    assert app_module._heuristic_dedicated("breast cancer", subs) == {
        "breastcancer", "breastcancerawareness",
    }
    # r/cancer is broader than the topic, never exempted by name.
    assert "cancer" not in app_module._heuristic_dedicated("breast cancer", subs)
    # Too-short squashed topics never match.
    assert app_module._heuristic_dedicated("ibs", ["ibs"]) == set()


def test_collector_filter_exempt_fake(monkeypatch):
    # With a topic filter, a dedicated (exempt) subreddit is exported in full
    # while the others are still keyword-filtered.
    monkeypatch.setenv("RTS_FAKE", "1")
    result = collector.collect(
        ["endometriosis", "PCOS"], 1_700_000_000, 1_700_200_000,
        keywords=["no-such-word"],
        filter_exempt={"endometriosis"},
    )
    by_sub = {}
    for p in result["posts"]:
        by_sub.setdefault(p["subreddit"], []).append(p)
    assert len(by_sub.get("endometriosis", [])) == 3   # unfiltered: all posts
    assert "PCOS" not in by_sub                        # filtered: nothing matches
    assert all(c["subreddit"] == "endometriosis" for c in result["comments"])


def test_collect_endpoint_dedicated_sub_unfiltered(client, monkeypatch):
    # Topic "synthetic": r/synthetic_support matches by name -> exported in
    # full (posts AND comments); r/endometriosis stays keyword-filtered, and
    # its comments (which never say "synthetic") are dropped.
    monkeypatch.setenv("RTS_FAKE", "1")
    resp = client.post(
        "/api/collect",
        json={
            "subreddits": ["synthetic_support", "endometriosis"],
            "start_date": "2023-11-01",
            "end_date": "2023-11-30",
            "topic": "synthetic",
        },
    )
    assert resp.status_code == 200
    assert int(resp.headers["X-Collect-Posts"]) == 6      # 3 unfiltered + 3 match
    assert int(resp.headers["X-Collect-Comments"]) == 5   # exempt sub only
    assert resp.headers.get("X-Collect-Unfiltered") == "synthetic_support"


def test_fallback_keywords():
    kws = app_module._fallback_keywords("Chemotherapy and hair loss")
    assert "chemotherapy and hair loss" in kws  # full phrase first
    assert "chemotherapy" in kws and "hair" in kws and "loss" in kws
    assert "and" not in kws  # stopword dropped
    assert app_module._fallback_keywords("") == []
    assert app_module._fallback_keywords("lupus") == ["lupus"]


def test_bundle_groups_by_subreddit():
    zip_bytes = bundle.build_zip_bytes(
        _stub_posts(),
        _stub_comments(),
        spreadsheet.build_workbook_bytes(_stub_posts(), _stub_comments()),
        "reddit_export_test",
    )
    zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
    assert "reddit_export_test.xlsx" in zf.namelist()
    posts = _read_ndjson_zst(zf.read("raw/endometriosis_posts.ndjson.zst"))
    assert [p["id"] for p in posts] == ["p1", "p2"]
    comments = _read_ndjson_zst(zf.read("raw/endometriosis_comments.ndjson.zst"))
    assert len(comments) == 3


def test_collector_arctic_shift_fallback(monkeypatch):
    # pullpush fails outright; arctic_shift serves one page then runs dry.
    monkeypatch.delenv("RTS_FAKE", raising=False)
    monkeypatch.setattr(collector, "REQUEST_SLEEP", 0)

    calls = []

    def fake_fetch(url, params):
        calls.append(url)
        if "pullpush.io" in url:
            raise RuntimeError("HTTP 502")
        if len(calls) <= 2:  # first arctic page has data, the next is empty
            return [{
                "id": "as1", "subreddit": "lupus", "created_utc": 1_700_000_100,
                "author": "arctic_user", "title": "From arctic_shift",
                "selftext": "fallback works", "score": 1, "num_comments": 0,
                "permalink": "/r/lupus/comments/as1/",
            }]
        return []

    monkeypatch.setattr(collector, "_fetch", fake_fetch)

    result = collector.collect(
        ["lupus"], 1_700_000_000, 1_700_001_000, include_comments=False,
    )

    assert [p["id"] for p in result["posts"]] == ["as1"]
    assert any("falling back to arctic_shift" in e for e in result["errors"])
    assert any("pullpush" in e for e in result["errors"])


def test_collector_arctic_shift_on_empty_pullpush(monkeypatch):
    # pullpush answers 200 with NO rows (its "up but empty index" mode, which
    # produced silently empty exports); arctic_shift must be consulted before
    # believing the window is empty.
    monkeypatch.delenv("RTS_FAKE", raising=False)
    monkeypatch.setattr(collector, "REQUEST_SLEEP", 0)

    arctic_pages = []

    def fake_fetch(url, params):
        if "pullpush.io" in url:
            return []
        arctic_pages.append(params)
        if len(arctic_pages) == 1:
            return [{
                "id": "as9", "subreddit": "lupus", "created_utc": 1_700_000_100,
                "author": "x", "title": "Only arctic has this",
                "selftext": "", "score": 2, "num_comments": 0,
                "permalink": "/r/lupus/comments/as9/",
            }]
        return []

    monkeypatch.setattr(collector, "_fetch", fake_fetch)

    result = collector.collect(
        ["lupus"], 1_700_000_000, 1_700_001_000, include_comments=False,
    )

    assert [p["id"] for p in result["posts"]] == ["as9"]
    assert result["errors"] == []  # empty-index fallback is not an error


def test_collector_time_budget_partial(monkeypatch):
    # An exhausted time budget must stop collection cleanly with a note per
    # sub/kind and keep whatever was already collected (here: nothing), never
    # raise -- this is what turns Vercel's FUNCTION_INVOCATION_TIMEOUT into a
    # partial export.
    monkeypatch.delenv("RTS_FAKE", raising=False)
    monkeypatch.setattr(collector, "TIME_BUDGET", 1e-9)  # expires immediately

    def fail_if_called(url, params):
        raise AssertionError("no network call should happen after the deadline")

    monkeypatch.setattr(collector, "_fetch", fail_if_called)

    result = collector.collect(["lupus"], 1_700_000_000, 1_700_001_000)

    assert result["posts"] == [] and result["comments"] == []
    assert len(result["errors"]) == 2  # posts + comments both noted
    assert all(collector._BUDGET_NOTE in e for e in result["errors"])


def test_collector_budget_is_split_fairly(monkeypatch):
    # A subreddit with endless posts must NOT starve the comments task: the
    # budget is sliced per (sub, kind), so posts get cut at their slice and
    # comments still run. (This is the "11k posts, 0 comments" regression.)
    monkeypatch.delenv("RTS_FAKE", raising=False)
    monkeypatch.setattr(collector, "REQUEST_SLEEP", 0.01)
    monkeypatch.setattr(collector, "TIME_BUDGET", 0.6)

    comment_pages = []

    def fake_fetch(url, params):
        import time as _t
        _t.sleep(0.02)
        if "comment" in url:
            comment_pages.append(1)
            if len(comment_pages) == 1:
                return [{
                    "id": "c1", "subreddit": "lupus", "created_utc": 500_000,
                    "author": "a", "body": "a comment", "score": 1,
                    "link_id": "t3_x", "parent_id": "t3_x", "permalink": "/c1",
                }]
            return []
        # Endless stream of posts, walking backwards in time forever.
        return [{
            "id": f"p{params['before']}", "subreddit": "lupus",
            "created_utc": int(params["before"]) - 1, "author": "a",
            "title": "endless", "selftext": "", "score": 1,
            "num_comments": 0, "permalink": "/p",
        }]

    monkeypatch.setattr(collector, "_fetch", fake_fetch)

    result = collector.collect(["lupus"], 0, 1_000_000)

    assert len(result["posts"]) > 0          # posts collected until their slice
    assert len(result["comments"]) == 1      # comments still got their turn
    assert any("post" in e and collector._BUDGET_NOTE in e for e in result["errors"])


def test_collector_keyword_filter_fake(monkeypatch):
    monkeypatch.setenv("RTS_FAKE", "1")
    result = collector.collect(
        ["endometriosis"], 1_700_000_000, 1_700_200_000,
        keywords=["post #1"],  # matches exactly one synthetic post title
    )
    assert len(result["posts"]) == 1
    assert "#1" in result["posts"][0]["title"]
    assert result["comments"] == []


def test_collect_endpoint_empty_subreddits_is_400(client, monkeypatch):
    monkeypatch.setenv("RTS_FAKE", "1")
    resp = client.post(
        "/api/collect",
        json={
            "subreddits": [],
            "start_date": "2023-11-01",
            "end_date": "2023-11-30",
        },
    )
    assert resp.status_code == 400
    assert "error" in resp.get_json()


def test_collect_endpoint_bad_date_is_400(client, monkeypatch):
    monkeypatch.setenv("RTS_FAKE", "1")
    resp = client.post(
        "/api/collect",
        json={
            "subreddits": ["endometriosis"],
            "start_date": "not-a-date",
            "end_date": "2023-11-30",
        },
    )
    assert resp.status_code == 400
    assert "error" in resp.get_json()


def test_subreddits_endpoint(client):
    resp = client.get("/api/subreddits")
    assert resp.status_code == 200
    data = resp.get_json()
    assert "themes" in data
    assert "Women's health" in data["themes"]
    assert "endometriosis" in data["themes"]["Women's health"]
    assert "pool" in data and "endometriosis" in data["pool"]
    assert "popular" in data and data["popular"]


def test_suggest_endpoint_fallback(client, monkeypatch):
    # RTS_FAKE=1 skips Mercury, so this exercises the static fallback offline.
    monkeypatch.setenv("RTS_FAKE", "1")
    resp = client.post("/api/suggest", json={"selected": ["endometriosis"]})
    assert resp.status_code == 200
    data = resp.get_json()
    assert data["source"] == "fallback"
    names = [s["name"].lower() for s in data["suggestions"]]
    assert names, "expected related suggestions"
    assert "endometriosis" not in names  # never suggest what's already selected
    # endometriosis is women's health -> should surface other women's-health subs
    assert any(n in names for n in ("pcos", "adenomyosis", "pmdd"))


def test_chat_endpoint_fallback(client, monkeypatch):
    # RTS_FAKE=1 skips Mercury -> exercises the keyword/theme fallback offline.
    monkeypatch.setenv("RTS_FAKE", "1")
    resp = client.post("/api/chat", json={"message": "lupus and joint pain"})
    assert resp.status_code == 200
    data = resp.get_json()
    names = [s.lower() for s in data["subreddits"]]
    assert "lupus" in names            # directly named
    assert len(names) > 1              # expanded to the autoimmune theme
    assert data["reply"]

    empty = client.post("/api/chat", json={"message": ""})
    assert empty.status_code == 200
    assert empty.get_json()["subreddits"] == []
