"""Excel workbook builder for reddit-to-spreadsheet.

Turns collected Post / Comment records (see the shared schema in collector.py)
into a multi-sheet .xlsx workbook with openpyxl.

Public interface:
    build_workbook_bytes(posts, comments) -> bytes
    build_workbook(posts, comments, out_path) -> str

Sheets produced:
    "Posts"    always (headers only when there are no posts)
    "Comments" only when there is at least one comment
    "Summary"  always -- per-subreddit stats, a TOTAL row, and a by-year block
"""

from __future__ import annotations

import io
from collections import defaultdict
from datetime import datetime, timezone

from openpyxl import Workbook
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# Excel's hard limit is 32767 chars per cell; stay comfortably under it.
_MAX_CELL_LEN = 32000

_HEADER_FONT = Font(bold=True)

# --- Posts sheet -----------------------------------------------------------
_POST_HEADERS = [
    "subreddit", "id", "created", "author", "title",
    "selftext", "score", "num_comments", "permalink",
]
_POST_WIDTHS = [18, 12, 20, 18, 50, 60, 8, 14, 40]

# --- Comments sheet --------------------------------------------------------
_COMMENT_HEADERS = [
    "subreddit", "id", "created", "author", "body",
    "score", "link_id", "parent_id", "permalink",
]
_COMMENT_WIDTHS = [18, 12, 20, 18, 60, 8, 14, 14, 40]

# --- Summary sheet ---------------------------------------------------------
_SUMMARY_HEADERS = [
    "subreddit", "posts", "comments", "unique_authors",
    "total_words", "earliest", "latest",
]
_SUMMARY_WIDTHS = [22, 8, 10, 16, 12, 22, 22]


def _iso(created_utc) -> str:
    """UTC ISO-8601 timestamp.

    Equivalent to the contract's datetime.utcfromtimestamp(ts).isoformat(),
    but written with a tz-aware datetime so it stays warning-free on 3.12+.
    """
    dt = datetime.fromtimestamp(int(created_utc), tz=timezone.utc)
    return dt.replace(tzinfo=None).isoformat()


def _year(created_utc) -> int:
    """Calendar year (UTC) of a unix timestamp."""
    return datetime.fromtimestamp(int(created_utc), tz=timezone.utc).year


def _trunc(value):
    """Make a value safe for an xlsx cell; pass non-strings through.

    Strings are scrubbed of control characters that openpyxl rejects with
    IllegalCharacterError (real Reddit text does contain them) and clamped to
    the cell length limit. Only the spreadsheet is scrubbed -- the raw
    .ndjson.zst export keeps the original text untouched.
    """
    if isinstance(value, str):
        value = ILLEGAL_CHARACTERS_RE.sub("", value)
        if len(value) > _MAX_CELL_LEN:
            value = value[:_MAX_CELL_LEN]
    return value


def _write_grid(ws, headers, rows, widths):
    """Write a bold, frozen header row plus data rows, then set column widths."""
    ws.append(headers)
    for cell in ws[1]:
        cell.font = _HEADER_FONT
    ws.freeze_panes = "A2"  # keep the header visible while scrolling
    for row in rows:
        ws.append([_trunc(v) for v in row])
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _post_row(p):
    return [
        p.get("subreddit", ""),
        p.get("id", ""),
        _iso(p.get("created_utc", 0)),
        p.get("author"),
        p.get("title", "") or "",
        p.get("selftext", "") or "",
        p.get("score", 0),
        p.get("num_comments", 0),
        p.get("permalink", ""),
    ]


def _comment_row(c):
    return [
        c.get("subreddit", ""),
        c.get("id", ""),
        _iso(c.get("created_utc", 0)),
        c.get("author"),
        c.get("body", "") or "",
        c.get("score", 0),
        c.get("link_id", ""),
        c.get("parent_id", ""),
        c.get("permalink", ""),
    ]


def _build_summary(ws, posts, comments):
    """Populate the Summary sheet: per-subreddit stats, TOTAL, and by-year block."""
    # Per-subreddit accumulators.
    stats = defaultdict(lambda: {
        "posts": 0, "comments": 0, "authors": set(),
        "words": 0, "earliest": None, "latest": None,
    })
    by_year = defaultdict(lambda: [0, 0])  # year -> [posts, comments]

    def note_time(bucket, ts):
        if bucket["earliest"] is None or ts < bucket["earliest"]:
            bucket["earliest"] = ts
        if bucket["latest"] is None or ts > bucket["latest"]:
            bucket["latest"] = ts

    for p in posts:
        b = stats[p.get("subreddit", "")]
        b["posts"] += 1
        author = p.get("author")
        if author:
            b["authors"].add(author)
        # total_words = whitespace-split word count of title + selftext.
        text = f"{p.get('title') or ''} {p.get('selftext') or ''}"
        b["words"] += len(text.split())
        ts = int(p.get("created_utc", 0))
        note_time(b, ts)
        by_year[_year(ts)][0] += 1

    for c in comments:
        b = stats[c.get("subreddit", "")]
        b["comments"] += 1
        author = c.get("author")
        if author:
            b["authors"].add(author)
        # total_words = whitespace-split word count of body.
        b["words"] += len(str(c.get("body") or "").split())
        ts = int(c.get("created_utc", 0))
        note_time(b, ts)
        by_year[_year(ts)][1] += 1

    # Header row (bold + frozen).
    ws.append(_SUMMARY_HEADERS)
    for cell in ws[1]:
        cell.font = _HEADER_FONT
    ws.freeze_panes = "A2"

    # One row per subreddit, plus running totals for the TOTAL row.
    tot_posts = tot_comments = tot_words = 0
    all_authors = set()
    overall_earliest = overall_latest = None

    for sub in sorted(stats):
        b = stats[sub]
        ws.append([
            _trunc(sub),
            b["posts"],
            b["comments"],
            len(b["authors"]),
            b["words"],
            _iso(b["earliest"]) if b["earliest"] is not None else "",
            _iso(b["latest"]) if b["latest"] is not None else "",
        ])
        tot_posts += b["posts"]
        tot_comments += b["comments"]
        tot_words += b["words"]
        all_authors |= b["authors"]
        if b["earliest"] is not None:
            overall_earliest = (b["earliest"] if overall_earliest is None
                                else min(overall_earliest, b["earliest"]))
        if b["latest"] is not None:
            overall_latest = (b["latest"] if overall_latest is None
                              else max(overall_latest, b["latest"]))

    # TOTAL row -- unique_authors is the distinct union across all subreddits.
    ws.append([
        "TOTAL",
        tot_posts,
        tot_comments,
        len(all_authors),
        tot_words,
        _iso(overall_earliest) if overall_earliest is not None else "",
        _iso(overall_latest) if overall_latest is not None else "",
    ])
    for cell in ws[ws.max_row]:
        cell.font = _HEADER_FONT

    # By-year block: blank spacer, a label, a bold header, then one row/year.
    ws.append([])
    ws.append(["By year"])
    ws.cell(row=ws.max_row, column=1).font = _HEADER_FONT
    ws.append(["year", "posts", "comments"])
    for cell in ws[ws.max_row]:
        cell.font = _HEADER_FONT
    for year in sorted(by_year):
        yp, yc = by_year[year]
        ws.append([year, yp, yc])

    for idx, width in enumerate(_SUMMARY_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def _build(posts, comments) -> Workbook:
    """Assemble the workbook. Empty input still yields valid header-only sheets."""
    posts = posts or []
    comments = comments or []

    wb = Workbook()

    ws_posts = wb.active
    ws_posts.title = "Posts"
    _write_grid(ws_posts, _POST_HEADERS, [_post_row(p) for p in posts], _POST_WIDTHS)

    if comments:  # only add the Comments sheet when there is something to show
        ws_comments = wb.create_sheet("Comments")
        _write_grid(ws_comments, _COMMENT_HEADERS,
                    [_comment_row(c) for c in comments], _COMMENT_WIDTHS)

    ws_summary = wb.create_sheet("Summary")
    _build_summary(ws_summary, posts, comments)

    return wb


def build_workbook_bytes(posts: list[dict], comments: list[dict]) -> bytes:
    """Build the workbook and return it as raw .xlsx bytes (for HTTP download)."""
    wb = _build(posts, comments)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_workbook(posts: list[dict], comments: list[dict], out_path: str) -> str:
    """Build the workbook, save it to out_path, and return that path."""
    wb = _build(posts, comments)
    wb.save(out_path)
    return out_path
